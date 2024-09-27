# ----------------- Imports ---------------------#
from tkinter import *
from tkinter import messagebox
from PIL import ImageTk, Image
from tkinter.ttk import Combobox
from matplotlib import *
import numpy as np
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from easygui import *
import queue

# Sets Graph info as None Initially
fig = None
ax = None
pie_chart_canvas = None


#-------------------------------------------- GAME CLASS ---------------------------------------------------------#
class Game:
    
    # ----------------- Initialisation ---------------------#
    def __init__(self, title, option1_score, option2_score, option3_score, option4_score, option5_score, option6_score, option7_score, option8_score, option9_score, option10_score, review, tags):
        self.title = title
        self.option1_score = int(option1_score)
        self.option2_score = int(option2_score)
        self.option3_score = int(option3_score)
        self.option4_score = int(option4_score)
        self.option5_score = int(option5_score)
        self.option6_score = int(option6_score)
        self.option7_score = int(option7_score)
        self.option8_score = int(option8_score)
        self.option9_score = int(option9_score)
        self.option10_score = int(option10_score)
        self.tags = tags
        self.review = review
        self.total_score = self.calculate_total_score()

    # ----------------- Method1: Calculate Total Score ---------------------#
    def calculate_total_score(self):

        return round(sum([self.option1_score, self.option2_score, self.option3_score, self.option4_score,
                    self.option5_score, self.option6_score, self.option7_score, self.option8_score,
                    self.option9_score, self.option10_score]) / 10,2)

#-------------------------------------------- GAME MANAGER CLASS ---------------------------------------------------------#
class GameManager:

    #---------- Initialisation --------#
    def __init__(self, file_path):
        self.file_path = file_path
        self.games = self.load_games()
        self.game_queue = queue.Queue()  

    #---------- Method1: Load Games --------#
    def load_games(self):
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            games = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) == 15:  # Adjusted to 15 to match all columns
                    game = Game(*row)
                    games.append(game)
                else:
                    print("Skipping Row")
            return games
        except FileNotFoundError:
            return []

    #---------- Method 2: Save Games in File --------#
    def save_games(self, game=None):

        # Loads Workbook
        try:

            wb = load_workbook(self.file_path)
            ws = wb.active

        except FileNotFoundError:
            
            #If Excel File is not created then it will create one
            wb = Workbook()
            ws = wb.active
            ws.append(["Title", "Total Score", "Performance", "Audio Elements", "Visual Elements", "Story", "Side Content", "Updates", "Originality", "Accessibility", "Enjoyment", "Replayability", "Review", "Tags"])

        # Shows that game is saves
        if game:
            print(f"Saving game: {game.title}")
            
            # Checks if row is filled or not
            next_row = ws.max_row + 1
            if ws.cell(row=next_row, column=1).value is not None:
                next_row += 1
            
            # Fills columns wth data
            ws.cell(row=next_row, column=1, value=game.title)
            ws.cell(row=next_row, column=2, value=game.total_score)
            ws.cell(row=next_row, column=3, value=game.option1_score)
            ws.cell(row=next_row, column=4, value=game.option2_score)
            ws.cell(row=next_row, column=5, value=game.option3_score)
            ws.cell(row=next_row, column=6, value=game.option4_score)
            ws.cell(row=next_row, column=7, value=game.option5_score)
            ws.cell(row=next_row, column=8, value=game.option6_score)
            ws.cell(row=next_row, column=9, value=game.option7_score)
            ws.cell(row=next_row, column=10, value=game.option8_score)
            ws.cell(row=next_row, column=11, value=game.option9_score)
            ws.cell(row=next_row, column=12, value=game.option10_score)
            ws.cell(row=next_row, column=13, value=game.review)
            ws.cell(row=next_row, column=14, value=game.tags)

        #Save Workbook
        wb.save(self.file_path)



    #---------- Method 3: Add Game to Excel File --------#
    def add_game(self, game):
        self.save_games(game)
        self.process_game_queue()

    #---------- Method4: To Process the Queue --------#
    def process_game_queue(self):
        while not self.game_queue.empty():
            game_to_save = self.game_queue.get()
            self.save_games(game_to_save)
            self.game_queue.task_done()  # Mark the task as done

#-------------------------------------------- Main Program ---------------------------------------------------------#
def main():

    # ----------------- Upload Image Function ---------------------#
    def uploading_image():
        
        #Updates the button with the new image
        def update_image():

            # Gets image path and then displays it where the button was
            global image_path
            image_path = output
            img = Image.open(image_path)
            img = img.resize((500, 500))
            img = ImageTk.PhotoImage(img)
            image_button.config(image=img, text="",width=500, height=500)
            image_button.pack_configure(expand=True)
            image_button.image = img  
            

        # Using EasyGui it displays and enterbox where users can input image_path
        text = "Enter Image URL"
        title = "Add Image"
        d_text = "Enter Here"
        output = enterbox(text, title, d_text)

        update_image()

    # ----------------- Display Graph Function ---------------------#
    def plot_pie_chart():

        # Gets globals
        global fig, ax, pie_chart_canvas, root

        # Get scores from the comboboxes
        scores = []
        for score_entry in score_entries:
            score = score_entry.get()

            if not score:
                messagebox.showinfo(title="Error", message="All score entries must be filled")
                return
            
            elif int(score) < 0 or int(score) > 10:
                messagebox.showinfo(title="Error", message="Scores cannot be above 10 or below 0")
                return

            scores.append(int(score_entry.get()))

        #Display Root
        root = Tk()
        root.title("Graph")
        root.geometry("1200x574")

        #Labels and Buttons
        global scoring_frame4
        scoring_frame4 = Frame(root, bg="gray", width=1200, height=574, padx=5, pady=5)
        scoring_frame4.pack(side="top", fill=BOTH)

        graph_btn = Button(scoring_frame4, text="Hide Graph", font=('Segoe UI', 20, "bold"), bg="orange", fg="black", height=1, command=hide_pie_chart)
        graph_btn.pack(side="top")

        mylabels = ["Performance", "Audio", "Visual", "Story", "Side Content", "Updates", "Originality", "Accessibility", "Enjoyment", "Replayability"]

        # Create new figure and axis for each plot
        fig, ax = plt.subplots()
        ax.pie(scores, autopct='%1.1f%%', labels=mylabels, startangle=90)
        ax.axis('equal')

        # Clears previous canvas if it exists
        if pie_chart_canvas:
            pie_chart_canvas.get_tk_widget().pack_forget()

        # Attach the new pie chart to the canvas
        pie_chart_canvas = FigureCanvasTkAgg(fig, scoring_frame4)
        pie_chart_canvas.draw()
        pie_chart_canvas.get_tk_widget().pack(side=TOP)

        root.mainloop()

    # ----------------- Hide Graph Function ---------------------#
    def hide_pie_chart():

        #Destroys root and resets pie chart
        global root, fig, ax, pie_chart_canvas
        if root is not None:
            root.destroy()
            root = None  
        fig = None  
        ax = None  
        pie_chart_canvas = None  

    # ----------------- Calculate Overall Score Function ---------------------#
    def calculate_average_scoring():
        
        #Initial Average Score
        average_score = 0

        #Gets scores from comboboxes
        for score_entry in score_entries:
            score = score_entry.get()

            if not score:
                messagebox.showinfo(title="Error", message="All score entries must be filled")
                return
            
            if int(score) < 0 or int(score) > 10:
                messagebox.showinfo(title="Error", message="Scores cannot be above 10 or below 0")
                return

            
            average_score += int(score_entry.get())


        #Calculates Average Score
        average_score = round((average_score/10),2)
        overall_game_rating.config(fg="white")
        
        #Displays background colour based off from score
        if average_score < 5:
            overall_game_rating.config(bg="red")
        
        elif average_score < 7.5:
            overall_game_rating.config(bg="#F48C10")
        
        else: 
            overall_game_rating.config(bg="green")

        # Configs all buttons 
        calculate_game_rating_btn.config(bg="light green")
        overall_game_rating.configure(text="{}".format(str(average_score)), bd=3,relief="solid")
        overall_game_rating.pack_configure(padx=(65,200))

    # ----------------- Searching for Game Function ---------------------#
    def search(event):

        #Gets whatever is typed
        typed = game_name.get()

        #If nothing is typed show all games
        if typed == '':
            game_name['values'] = list_of_games  

        else:
            
            #If types then it filters games
            filtered_games = [game for game in list_of_games if typed.lower() in game.lower()]
            game_name['values'] = filtered_games  
    
    # ----------------- Complete Game Function ---------------------#
    def complete_game():

        #Checks Image Path
        global image_path
        try:
            if image_path == None:
                image_path = ""
        except NameError:
            image_path = ""

        
        tags = image_path

        #Check Review Box
        review = text_box.get('1.0',END)

        # Checks length of Review
        if len(review) < 31 or len(review) > 500 :
            messagebox.showerror("Error","Review must be between 30 and 500 characters in length")
            return

        #Establishes Game Manager
        game_manager = GameManager("games_review3.xlsx")

        #Gets all scores from score_entries
        for score_entry in score_entries:
            score = score_entry.get()

            if not score:
                messagebox.showinfo(title="Error", message="All score entries must be filled")
                return

            elif int(score) < 0 or int(score) > 10:
                messagebox.showinfo(title="Error", message="Scores cannot be above 10 or below 0")
                return
            

        # Check if all comboboxes are filled
        try:
            if game_name.get() == "":
                messagebox.showinfo(title="Error", message="Game Name cannot be empty")
                return
            
            elif game_name.get() not in list_of_games:
                messagebox.showinfo(title="Error", message="Game Name Not Found")
                return

            for score_entry in score_entries:
                if score_entry.get() == "":
                    messagebox.showinfo(title="Error", message="All score entries must be filled")
                    return

        except ValueError:
            messagebox.showinfo(title="Error", message="Make sure all entries are filled correctly")
            return

        # Create and add the game to the manager
        new_game = Game(
            game_name.get(),
            score_entries[0].get(), score_entries[1].get(), score_entries[2].get(), score_entries[3].get(),
            score_entries[4].get(), score_entries[5].get(), score_entries[6].get(), score_entries[7].get(),
            score_entries[8].get(), score_entries[9].get(),
            review, tags
        )
        
        #Puts Game in game manager to be stored and then puts user back to home page
        game_manager.add_game(new_game)
        messagebox.showinfo(title="Success", message="Game completed successfully!")
        window.destroy()
        from TrueIteration3MainPage import mainloop
        



    # window setup
    window = Tk() 
    window.title("Game Page")
    window.geometry("1920x900")

    pxl = Image.open('pixel.png')
    pxl = ImageTk.PhotoImage(pxl)

    # Task Bar
    task_bar_frame = Frame(window, bg="#236A43", width=1920, height=100)
    task_bar_frame.pack(side="top")

    task_bar_btn = Button(task_bar_frame, fg="white", compound="center", text="Game Critic Program", bg="#236A43", image=pxl, width=1920, height=100, font=('Segoe UI', 50, "bold"), border=0)
    task_bar_btn.pack(side="top")

    # Game Frame with Scrollbar
    game_frame = Frame(window, height=1277, width=1920, bg="light gray")
    game_frame.pack(side="top", fill="both", expand=True)

    # Create a canvas
    canvas = Canvas(game_frame, bg="light gray")
    canvas.pack(side="left", fill="both", expand=True)

    # Add a scrollbar to the canvas
    scrollbar = Scrollbar(game_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    # Configure the canvas
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    # Create another frame inside the canvas
    inner_frame = Frame(canvas, bg="light gray")
    canvas.create_window((0, 0), window=inner_frame, anchor="nw")

    # Add Game Title
    title_frame = Frame(inner_frame, bg="light gray")
    title_frame.pack(side="top", fill="x", pady=10)

    game_name_lbl = Label(title_frame, text="Game Name: ", font=('Segoe UI', 30, "bold"), fg="black", bg="light gray", compound="center")
    game_name_lbl.pack(side="left",padx=(290,115))

    from list_of_games import list_of_games

    game_name = Combobox(title_frame, width=38, font=('Segoe UI', 30, "bold"), values=list_of_games)
    game_name.pack(side="left")

    game_name.bind("<KeyRelease>", search)

    


    # ScoreBoard
    scoring_frame = Frame(inner_frame, width=1700, height=584, bg="gray")
    scoring_frame.pack(side="top", pady=(0,20), fill="y", expand=True,padx=(130,0))

    # ScoreBoard Inner Frames

    black_frame3 = Frame(scoring_frame, width=800, height=584, bg="black", border=2)
    black_frame3.pack(side="left", fill="y")

    scoring_frame3 = Frame(black_frame3, bg="gray", width=800, height=574, padx=5, pady=5)
    scoring_frame3.pack(side="left", fill="y")

    black_frame2 = Frame(scoring_frame, width=1200, height=584, bg="black", border=2)
    black_frame2.pack(side="left", fill="y")

    scoring_frame2 = Frame(black_frame2, bg="gray", width=1200, height=574, padx=5)
    scoring_frame2.pack(side="left", fill="y")

    # Info in Inner Frames

    lbl_frame = Frame(scoring_frame2, width=1200, height=0, bg="gray")
    lbl_frame.pack(side="top", fill="x", expand=TRUE)

    game_stat_lbl = Label(lbl_frame, justify="left",text="In-Game Scoring", font=('Segoe UI', 30, "bold"), bg="gray", fg="black")
    game_stat_lbl.pack(side="left", padx=(50,0))

    calculate_game_rating_btn = Button(lbl_frame,justify="right",text="Calculate Score",font=('Segoe UI', 20, "bold"),bg="#FFC300",fg="black",command=calculate_average_scoring, bd=3, relief="solid")
    calculate_game_rating_btn.pack(side="right",padx=(0,30),pady=7)

    options = [0,1,2,3,4,5,6,7,8,9,10]
    lbl_frames = []
    global score_entries
    score_entries = []

    for i in range(5):
        lbl_frame = Frame(scoring_frame2, width=1200, height=0, bg="gray")
        lbl_frame.pack(side="top", fill="x", expand=TRUE, pady=0)
        lbl_frames.append(lbl_frame)
    
    lbl_frame = Frame(scoring_frame2, width=1200, height=0, bg="gray",bd=3, relief="solid")
    lbl_frame.pack(side="top", fill="x", expand=TRUE, pady=0)
    lbl_frames.append(lbl_frame)

    performance_lbl = Label(lbl_frames[0], text="Performance:     ", font=('Segoe UI', 24, "bold"), bg="gray", fg="black")
    performance_lbl.pack(side="left", anchor=W, padx=(100, 20))

    audio_lbl = Label(lbl_frames[1], text="Audio:                ", font=('Segoe UI', 24, "bold"), bg="gray", fg="black")
    audio_lbl.pack(side="left", fill="x", padx=(100, 20))

    visual_lbl = Label(lbl_frames[2], text="Visual:                ", font=('Segoe UI', 24, "bold"), bg="gray", fg="black")
    visual_lbl.pack(side="left", anchor=W, padx=(100, 20))

    story_lbl = Label(lbl_frames[3], text="Story:                 ", font=('Segoe UI', 24, "bold"), bg="gray", fg="black")
    story_lbl.pack(side="left", anchor=W, padx=(100, 20))

    Side_story_lbl = Label(lbl_frames[4], text="   Side Content:     ", font=('Segoe UI', 24, "bold"), bg="gray", fg="black")
    Side_story_lbl.pack(side="left", anchor=W, padx=(70, 20))

    overall_game_rating_lbl = Label(lbl_frames[5],text="Game Rating", font=('Segoe UI', 30, "bold"), bg="gray", fg="black")
    overall_game_rating_lbl.pack(side="left",padx=(50,0))

    overall_game_rating = Label(lbl_frames[5],text="X", font=('Segoe UI', 25, "bold"), bg="gray", fg="black")
    overall_game_rating.pack(side="left",padx=(80,200),ipadx=(10))

    display_graph_btn = Button(lbl_frames[5],text="Display Graph", font=('Segoe UI', 20, "bold"), bg="#29BFCD",fg="black",command=plot_pie_chart, bd=3, relief="solid")
    display_graph_btn.pack(side="right",padx=(5,30),pady=7)

    for i in range(5):
        score_entry = Combobox(lbl_frames[i], values=options, font=('Segoe UI', 20, "bold"), width=3, justify=CENTER)
        score_entry.pack(side="left", anchor=W)
        score_entries.append(score_entry)

    

    Updates_lbl = Label(lbl_frames[0], text="Updates:", font=('Segoe UI', 24, "bold"), bg="gray", fg="black")
    Updates_lbl.pack(side="left", anchor=W, padx=(100,105))

    originality_lbl = Label(lbl_frames[1], text="Originality:", font=('Segoe UI', 24, "bold"), bg="gray", fg="black")
    originality_lbl.pack(side="left", anchor=W, padx=(100, 75))

    accessibility_lbl = Label(lbl_frames[2], text="Accessibility:", font=('Segoe UI', 24, "bold"), bg="gray", fg="black")
    accessibility_lbl.pack(side="left", anchor=W, padx=(100, 50))

    enjoyment_lbl = Label(lbl_frames[3], text="Enjoyment:", font=('Segoe UI', 24, "bold"), bg="gray", fg="black")
    enjoyment_lbl.pack(side="left", anchor=W, padx=(100, 75))

    replayability_lbl = Label(lbl_frames[4], text="Replayability:", font=('Segoe UI', 24, "bold"), bg="gray", fg="black")
    replayability_lbl.pack(side="left", anchor=W, padx=(100, 40))

    for i in range(5):
        score_entry = Combobox(lbl_frames[i], values=options, font=('Segoe UI', 20, "bold"), width=3, justify=CENTER)
        score_entry.pack(side="right", anchor=W,padx=(0,30))
        score_entries.append(score_entry)

    image_button = Button(scoring_frame3,text=" + Add Image", font=('Segoe UI', 31, "bold"),bg="light gray",fg="black",width=20,height=8,command=uploading_image)
    image_button.pack(side="top")

    
    completion_frame = Frame(inner_frame,width=1700, height=100, bg="light gray")
    completion_frame.pack(side="top")

    review_comments = Label(completion_frame, fg="black", bg="pink", justify=LEFT, text="    Review Comments(*required) \n * Minimum of 30 characters, Maximum of 500 characters  ", font=('Segoe UI', 24, "bold"),bd=3,relief="solid")
    review_comments.pack(side="left",padx=(80,50),ipady=8)
    
    # Complete Button
    complete_btn = Button(completion_frame, fg="black", compound="center", text="Save &\nSubmit", bg="#1EFF83", image=pxl, width=300, height=100, font=('Segoe UI', 24, "bold"), border=0, command=complete_game,bd=3,relief="solid")
    complete_btn.pack(side="right")

    review_frame = Frame(inner_frame, width=40,height=40,bg="light gray")
    review_frame.pack(side="left",pady=20,padx=(100,0))
    

    text_box = Text(review_frame, width=205,height=10,font=('Segoe UI', 10, "bold"),bd=3)
    text_box.pack(side="left",padx=(20,0))


    


    window.mainloop()
