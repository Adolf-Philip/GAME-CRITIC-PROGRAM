from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

class Game:
    def __init__(self, name, title, shrt_des, total_score, option1_score, option2_score, option3_score, option4_score, review):
        self.name = name
        self.title = title
        self.short_description = shrt_des
        self.total_score = total_score
        self.option1_score = option1_score
        self.option2_score = option2_score
        self.option3_score = option3_score
        self.option4_score = option4_score
        self.review = review
    
    def calculate_overall_score(self):
        overall_score = ((self.total_score + self.option1_score + self.option2_score + self.option3_score + self.option4_score) / 1.5)
        return overall_score

    def to_dict(self):
        return {
            "name": self.name,
            "title": self.title,
            "short_description": self.short_description,
            "total_score": self.total_score,
            "option1_score": self.option1_score,
            "option2_score": self.option2_score,
            "option3_score": self.option3_score,
            "option4_score": self.option4_score,
            "review": self.review
        }

    def from_dict(data):
        return Game(
            data["name"], data["title"], data["short_description"], data["total_score"],
            data["option1_score"], data["option2_score"], data["option3_score"], data["option4_score"], data["review"]
        )

class GameManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.games = self.load_games()

    def load_games(self):
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            games = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                game = Game(*row)
                games.append(game)
            return games
        except FileNotFoundError:
            return []

    def save_games(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Title", "Short Description", "Total Score", "Performance", "Visual Elements", "Audio Elements", "Gameplay", "Review"])
        for game in self.games:
            ws.append([
                game.name, game.title, game.short_description, game.total_score,
                game.option1_score, game.option2_score, game.option3_score, game.option4_score, game.review
            ])
        wb.save(self.file_path)

    def add_game(self, game):
        self.games.append(game)
        self.save_games()

    def edit_game(self, name, title, new_game):
        for idx, game in enumerate(self.games):
            if game.name == name and game.title == title:
                self.games[idx] = new_game
                self.save_games()
                return True
        return False

    def delete_game(self, name, title):
        self.games = [game for game in self.games if not (game.name == name and game.title == title)]
        self.save_games()

    def get_games_by_name(self, name):
        return [game for game in self.games if game.name == name]

def get_valid_score(prompt, min_val=0, max_val=10):
    while True:
        try:
            score = int(input(prompt))
            if score < min_val or score > max_val:
                raise ValueError
            return score
        except ValueError:
            print(f"Error: Enter a valid integer between {min_val} and {max_val}.")

def get_valid_string(prompt, required=True):
    while True:
        value = input(prompt).strip()
        if required and not value:
            print("Error: This field is required.")
        else:
            return value

def get_valid_review(prompt, keywords, min_length=72):
    while True:
        review = input(prompt).strip()
        if len(review) < min_length:
            print(f"Error: Review must be at least {min_length} characters.")
        elif not all(keyword in review for keyword in keywords):
            print(f"Error: Review must include the keywords: {', '.join(keywords)}")
        else:
            return review

def add_game(manager, name):
    title = get_valid_string("Enter Title of Game: ")
    shrt_des = get_valid_string("Enter short description of game: ")
    option1_score = get_valid_score("Enter score for Performance (out of 10): ")
    option2_score = get_valid_score("Enter score for Visual Elements (out of 10): ")
    option3_score = get_valid_score("Enter score for Audio Elements (out of 10): ")
    option4_score = get_valid_score("Enter score for Gameplay (out of 10): ")
    total_score = get_valid_score("Enter total score for the game (out of 100): ", 0, 100)

    print("Requirements for Review:\n\n Review must contain the following: \n- Include reasons for selections of score (key words: Performance, Visual Elements, Audio Elements, Gameplay, Overall Score)\n- Must include at least 72 characters minimum")
    review = get_valid_review("Enter Review of the Game (*required): ", ["Performance", "Visual Elements", "Audio Elements", "Gameplay", "Overall Score"])

    game = Game(name, title, shrt_des, total_score, option1_score, option2_score, option3_score, option4_score, review)
    manager.add_game(game)
    print(f"Overall score: {game.calculate_overall_score():.2f}")

def edit_game(manager, name):
    title = get_valid_string("Enter the title of the game you want to edit: ")
    old_game = next((g for g in manager.games if g.name == name and g.title == title), None)
    if not old_game:
        print("Error: Game not found.")
        return

    shrt_des = get_valid_string(f"Enter new short description of game ({old_game.short_description}): ", required=False)
    option1_score = get_valid_score(f"Enter new score for Performance ({old_game.option1_score}): ")
    option2_score = get_valid_score(f"Enter new score for Visual Elements ({old_game.option2_score}): ")
    option3_score = get_valid_score(f"Enter new score for Audio Elements ({old_game.option3_score}): ")
    option4_score = get_valid_score(f"Enter new score for Gameplay ({old_game.option4_score}): ")
    total_score = get_valid_score(f"Enter new total score for the game ({old_game.total_score}): ", 0, 100)
    review = get_valid_string(f"Enter new Review of the Game ({old_game.review}): ", required=False)

    new_game = Game(
        name, title,
        shrt_des or old_game.short_description,
        total_score, option1_score, option2_score, option3_score, option4_score,
        review or old_game.review
    )

    if manager.edit_game(name, title, new_game):
        print("Game updated successfully!")
        print(f"Overall score: {new_game.calculate_overall_score():.2f}")
    else:
        print("Error: Game could not be updated.")

def delete_game(manager, name):
    title = get_valid_string("Enter the title of the game you want to delete: ")
    manager.delete_game(name, title)
    print("Game deleted successfully!")

def display_games(manager, name):
    user_games = manager.get_games_by_name(name)
    if not user_games:
        print("No games found for this user.")
    for game in user_games:
        print(f"Title: {game.title}")
        print(f"Description: {game.short_description}")
        print(f"Total Score: {game.total_score}")
        print(f"Performance: {game.option1_score}")
        print(f"Visual Elements: {game.option2_score}")
        print(f"Audio Elements: {game.option3_score}")
        print(f"Gameplay: {game.option4_score}")
        print("Review:")
        print(game.review)
        print("\n")

def main():
    print("Welcome to The Game Critic Program \n")

    name = get_valid_string("Enter Name: ")

    manager = GameManager("games.xlsx")

    while True:
        try:
            choice = int(input("1. Add game \n2. Edit Game \n3. Delete Game \n4. Display Games \n5. End\nEnter Number: "))
        except ValueError:
            print("Invalid choice. Please enter a number between 1 and 5.")
            continue
        
        if choice == 1:
            add_game(manager, name)
        elif choice == 2:
            edit_game(manager, name)
        elif choice == 3:
            delete_game(manager, name)
        elif choice == 4:
            display_games(manager, name)
        elif choice == 5:
            print("Thank you for using The Game Critic Program. Goodbye!")
            break
        else:
            print("Invalid choice. Please enter a number between 1 and 5.")

if __name__ == "__main__":
    main()

