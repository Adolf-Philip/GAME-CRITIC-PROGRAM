
# ----------------- Imports ---------------------#
from tkinter import *
from PIL import ImageTk, Image
from openpyxl import load_workbook
from tkinter import ttk
from tkinter.ttk import Combobox



#------------------------------------------------ Main Function -----------------------------------------------------#
def main():


    # ----------------- Import Sign in and Sign up ---------------------#
    from TrueIteration3SignInLogIn import username


    # ----------------- Add Games ---------------------#
    def add_game():

        # Destroys root and brings them to game page
        root.destroy()
        import TrueIteration3GamePage
        TrueIteration3GamePage.main()

    # ----------------- Logout Function ---------------------#
    def logout():

        # Destroys root and then brings them back to sign in page
        root.destroy()
        from TrueIteration3SignInLogIn import login_page

    # ----------------- Delete Game Function ---------------------#
    def delete_game(game_title):

        # Load the workbook
        file_path = 'games_review3.xlsx'
        workbook = load_workbook(file_path)
        sheet = workbook.active

        #Chooses row to delete based off the game title
        row_to_delete = None
        for row in range(2, sheet.max_row + 1):  
            if sheet[f'A{row}'].value == game_title:
                row_to_delete = row
                break
        
        #Deletes the row
        if row_to_delete:
            sheet.delete_rows(row_to_delete)
            workbook.save(file_path)
            workbook.close()

        # Refreshes the page
        root.destroy()
        main()  
        
    # ----------------- Add Review Function ---------------------#
    def show_review(game_title):

        #Creates Window
        window = Tk()
        window.title("Review Frame")
        window.geometry("500x500")

        # Load the workbook
        file_path = 'games_review3.xlsx'
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Gets the row where the game title is located
        review = ""
        for row in range(2, sheet.max_row + 1): 
            if sheet[f'A{row}'].value == game_title:
                review = sheet[f'M{row}'].value
                break

        # Create a scrollable frame for reviews
        canvas = Canvas(window)
        scrollbar = Scrollbar(window, orient="vertical", command=canvas.yview)
        scrollable_frame = Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        # Place Labels and Canvases
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        review_header = Label(scrollable_frame, text="Review Comments", justify=CENTER, font=('Segoe UI', 31, "bold"), fg="black")
        review_header.pack(side=TOP, pady=(0, 20))

        review_text = Label(scrollable_frame, text=f"{review}", justify=LEFT, font=('Segoe UI', 12), wraplength=450)
        review_text.pack(side=TOP, padx=20, pady=0)
        
        canvas.pack(side=LEFT, fill="both", expand=True)
        scrollbar.pack(side=RIGHT, fill="y")

        window.mainloop()
        
    # ----------------- Show Scoring Function ---------------------#
    def show_scoring(game_title):
        
        # Creates Window
        window = Tk()
        window.title("Scoring Frame")
        window.geometry("1650x680")
        window.config(bg="black")

        # Load the workbook
        file_path = 'games_review3.xlsx'
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Get the row where the game title is located and all relevant information
        for row in range(2, sheet.max_row + 1): 
            if sheet[f'A{row}'].value == game_title:
                total_score = sheet[f'B{row}'].value
                performance = sheet[f'C{row}'].value
                audio = sheet[f'D{row}'].value
                visual = sheet[f'E{row}'].value
                story = sheet[f'F{row}'].value
                side_content = sheet[f"G{row}"].value
                updates = sheet[f'H{row}'].value
                originality = sheet[f'I{row}'].value
                accessibility = sheet[f'J{row}'].value
                enjoyment = sheet[f'K{row}'].value
                replayability = sheet[f'L{row}'].value
                break
        
        
        # ScoreBoard
        scoring_frame = Frame(window, width=1700, height=584, bg="gray")
        scoring_frame.pack(side="top", pady=(0,20), fill="y", expand=True)

        # ScoreBoard Inner Frames
        black_frame1 = Frame(scoring_frame, width=80, height=584, bg="black", border=2)
        black_frame1.pack(side="left", fill="y")

        scoring_frame1 = Frame(black_frame1, bg="gray", width=70, height=574, padx=5, pady=5)
        scoring_frame1.pack(side="left", fill="y")

        black_frame2 = Frame(scoring_frame, width=1200, height=584, bg="black", border=2)
        black_frame2.pack(side="left", fill="y")

        scoring_frame2 = Frame(black_frame2, bg="gray", width=1200, height=574, padx=5, pady=5)
        scoring_frame2.pack(side="left", fill="y")

        black_frame3 = Frame(scoring_frame, width=430, height=584, bg="black", border=2)
        black_frame3.pack(side="left", fill="y")

        scoring_frame3 = Frame(black_frame3, bg="gray", width=430, height=574, padx=5, pady=5)
        scoring_frame3.pack(side="left", fill="y")

        # Info in Inner Frames
        statistics_lbl = Label(scoring_frame1, text="S\nT\nA\nT\nI\nS\nT\nI\nC\nS", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        statistics_lbl.pack(side="top", fill="y", expand=True)

        game_stat_lbl = Label(scoring_frame2, text="                                     In-Game Scoring:                                    ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        game_stat_lbl.pack(side="top", fill="y")

        options = [1,2,3,4,5,6,7,8,9,10]
        lbl_frames = []
        global score_entries
        score_entries = []

        for i in range(5):
            lbl_frame = Frame(scoring_frame2, width=1200, height=70, bg="gray")
            lbl_frame.pack(side="top", fill="x", expand=TRUE, pady=20)
            lbl_frames.append(lbl_frame)

        performance_lbl = Label(lbl_frames[0], text="Performance:     ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        performance_lbl.pack(side="left", anchor=W, padx=(100, 20))

        audio_lbl = Label(lbl_frames[1], text="Audio:                 ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        audio_lbl.pack(side="left", fill="x", padx=(100, 20))

        visual_lbl = Label(lbl_frames[2], text="Visual:                 ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        visual_lbl.pack(side="left", anchor=W, padx=(100, 20))

        story_lbl = Label(lbl_frames[3], text="Story:                  ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        story_lbl.pack(side="left", anchor=W, padx=(100, 20))

        Side_story_lbl = Label(lbl_frames[4], text="   Side Content:     ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        Side_story_lbl.pack(side="left", anchor=W, padx=(70, 20))

        score_data = [performance, audio, visual, story, side_content]
        for i in range(5):
            score_entry = Combobox(lbl_frames[i], values=options, font=('Segoe UI', 20, "bold"), width=3, justify=CENTER, state="disabled")
            score_entry.set(score_data[i])  # Set value from Excel
            score_entry.pack(side="left", anchor=W)
            score_entries.append(score_entry)

        

        Updates_lbl = Label(lbl_frames[0], text="    Updates:        ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        Updates_lbl.pack(side="left", anchor=W, padx=(100, 20))

        originality_lbl = Label(lbl_frames[1], text="  Originality:     ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        originality_lbl.pack(side="left", anchor=W, padx=(100, 20))

        accessibility_lbl = Label(lbl_frames[2], text="          Accessibility:   ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        accessibility_lbl.pack(side="left", anchor=W, padx=(0, 20))

        enjoyment_lbl = Label(lbl_frames[3], text="           Enjoyment:     ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        enjoyment_lbl.pack(side="left", anchor=W, padx=(0, 20))

        replayability_lbl = Label(lbl_frames[4], text="         Replayability:   ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        replayability_lbl.pack(side="left", anchor=W, padx=(0, 20))

        score_data_2 = [updates, originality, accessibility, enjoyment, replayability]
        for i in range(5):
            score_entry = Combobox(lbl_frames[i], values=options, font=('Segoe UI', 20, "bold"), width=3, justify=CENTER, state="disabled")
            score_entry.set(score_data_2[i])  # Set value from Excel
            score_entry.pack(side="left", anchor=W)
            score_entries.append(score_entry)

        
        # Overall Score and Average Score
        game_overall_lbl = Label(scoring_frame3, text="       Overall Score:        ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        game_overall_lbl.pack(side="top", fill="y")

        your_overall_score_lbl = Label(scoring_frame3, text=f"{(total_score * 10)}", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        your_overall_score_lbl.pack(side="top", pady=(50, 0))

        seperator_lbl = Label(scoring_frame3, text="-----------", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        seperator_lbl.pack(side="top", pady=20)

        total_score_lbl = Label(scoring_frame3, text="100", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        total_score_lbl.pack(side="top")

        average_score_lbl = Label(scoring_frame3, text="\n Average Score: ", font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        average_score_lbl.pack(side="top")

        calculated_average_score_lbl = Label(scoring_frame3, text="{}".format(round(total_score,2)), font=('Segoe UI', 31, "bold"), bg="gray", fg="black")
        calculated_average_score_lbl.pack(side="top")

    # ----------------- Apply Colour Function ---------------------#
    def apply_color(new_color):

        #Changes based on background colou
        if new_color == "#3299cc":
            my_reviews_btn.config(bg="white")
        
        elif new_color != "#3299cc":
            my_reviews_btn.config(bg="#29BFCD")
        
        # Update all frames and buttons with the new color
        taskbar_frame.config(bg=new_color)
        image_frame.config(bg=new_color)
        side_bar_options_frame.config(bg=new_color)
        new_releases_btn.config(bg=new_color)
        calendar_btn.config(bg=new_color)
        saved_btn.config(bg=new_color)
        logout_btn.config(bg=new_color)
        settings_btn.config(bg=new_color)
        separator.config(bg=new_color)
        separator_green.config(bg=new_color)
        separator_green2.config(bg=new_color)
        separator_green3.config(bg=new_color)
        separator_darkgreen.config(bg=new_color)
        logo.config(background=new_color)
        task_bar_label.config(bg=new_color)

    # ----------------- Setting Function ---------------------#
    def settings():
        
        #Creates Window
        window = Tk()
        window.title("Settings")
        window.geometry("300x300")

        colour_lbl = Label(window, font=('Segoe UI', 30, "bold"), fg="black", text="SET COLOUR:")
        colour_lbl.pack(side=TOP, pady=10)

        # Tkinter string variable able to store the selected value
        selected_color = StringVar(window, "1")

        # Dictionary of colors for the radio buttons
        values = {"Green": "#236A43",
                  "Blue": "#3299cc",
                  "Red": "#8A0303",
                  "Orange": "#FF9900"}

        # Create radio buttons for each color option
        for (text, color_code) in values.items():
            Radiobutton(window, text=text, variable=selected_color, value=color_code).pack(side=TOP, ipady=5)

        # Button to apply the selected color
        apply_btn = Button(window, text="Apply", command=lambda: apply_color(selected_color.get()))
        apply_btn.pack(pady=20)

        window.mainloop()

    # ----------------- Get Colour Function ---------------------#
    def get_color(average_score):

        #Depending on average score set the background colour
        if average_score < 5:
            return "red"
        elif 5 <= average_score <= 7.5:
            return "orange"
        else:
            return "green"


    # ----------------- Main Code ---------------------#
    root = Tk()
    root.title("Game Critic Starting Page")
    root.geometry("1500x1800")

    #Profile images
    root.photo1 = ImageTk.PhotoImage(Image.open("profilepic.png").resize((255, 218)))
    root.photo2 = ImageTk.PhotoImage(Image.open("logo.png").resize((120, 120)))

    #------------------------------------------ FRAMES ---------------------------------------------#
    taskbar_frame = Frame(root, bg="#236A43", height="50px", width="1500px")
    taskbar_frame.place(x=0,y=0)

    main_frame = Frame(root, width="1082px", height="1200px", bg="#7D7D7D", highlightthickness=0, highlightbackground="black", highlightcolor="black")
    main_frame.place(x=265,y=130)

    image_frame = Frame(root,bg="#236A43", height="100px", width="200px")
    image_frame.place(x=0,y=0)

    header_frame = Frame(root, bg="#232323", height="50px", width="1500px",highlightthickness=0, highlightbackground="black", highlightcolor="black")
    header_frame.place(x=265,y=65)

    sidebar_frame = Frame(root, bg="#232323", height="1000px", width="200px", highlightthickness=0, highlightbackground="black", highlightcolor="black")
    sidebar_frame.place(x=0,y=130)

    side_bar_options_frame = Frame(root, bg="#236A43", height="190px", width="190px", highlightthickness=2, highlightbackground="black", highlightcolor="black")
    side_bar_options_frame.place(x=7,y=390)

    #------------------------------------------  INNER FRAMES ---------------------------------------------#

    gl_frames = []
    xx = 35

    for i in range(4):
        game_library_frames = Frame(main_frame, width="200px", height="250px", bg="#9A9A9A", highlightthickness=2, highlightbackground="black", highlightcolor="black")
        game_library_frames.place(x=xx, y=10)
        xx += 300
        gl_frames.append(game_library_frames)

    xx = 35
    for i in range(4):
        game_library_frames = Frame(main_frame, width="200px", height="250px", bg="#9A9A9A", highlightthickness=2, highlightbackground="black", highlightcolor="black")
        game_library_frames.place(x=xx, y=405)
        xx += 300
        gl_frames.append(game_library_frames)

    separator = Frame(main_frame, width="1500px", height="40px", bg="#232323", highlightthickness=2, highlightbackground="black", highlightcolor="black")
    separator.place(x=0, y=347)

    #----------------------------------------------GET STORED GAMES------------------------------------------------------------#

    # Load the workbook
    file_path = 'games_review3.xlsx'
    workbook = load_workbook(file_path)

    # Get the game titles
    sheet = workbook.active
    a_values = [sheet[f'A{i}'].value for i in range(2, 10)] 

    # Displays 8 frames with stored games
    for i in range(8):
        frame = gl_frames[i]

        if i < len(a_values):
            
            #Text is value in sheet
            text = a_values[i] if a_values[i] else "Add Game [+]"
        else:
            # Text is "Add Game[+]"
            text = "Add Game [+]"

        if text == "Add Game [+]":

            # Creates a button for adding a game
            btn = Button(frame, width=18, text=text, bg="#1EFF83", fg="black", font=('Segoe UI', 18, "bold"), command=add_game)
            btn.place(x=0, y=0)

        else:

            # Creates a button for adding a game
            lbl = Label(frame, text=text, bg="#9A9A9A", font=('Segoe UI', 13, "bold"))
            lbl.place(x=5, y=5)

            # Retrieve the average score from column B for the corresponding game
            average_score = sheet[f'B{i+2}'].value if sheet[f'B{i+2}'].value else "No Rating"
            image_path = sheet[f'N{i+2}'].value if sheet[f'N{i+2}'].value else "NoImage.png"

            # checks if average score is an integer or float
            if isinstance(average_score, (int, float)):

                # Get background color based on the score
                background_color = get_color(average_score)

            else:

                # Default background if no rating
                background_color = "gray"  

            # Gets Image from Image Path
            img = Image.open(image_path)
            img = img.resize((235, 150))
            img = ImageTk.PhotoImage(img)
            image_button = Button(frame, image=img, text="",width=235, height=150)
            image_button.place(x=10,y=50)
            image_button.image = img  

            # Display the average score in the label
            overall_game_rating = Label(frame, text=f" {average_score} ",justify=CENTER, font=('Segoe UI', 30, "bold"), width=3,height=1,bg=background_color, fg="black" , bd=3, relief = "solid")
            overall_game_rating.place(x=18, y=232)

            delete_btn = Button(frame, width=2, text=" X ", bg="red", fg="black", font=('Segoe UI', 13, "bold"), command=lambda title=text: delete_game(title))
            delete_btn.place(x=230, y=2)

            show_scoring_btn = Button(frame, width=11, text="Show Scoring", bg="#29BFCD", fg="black", font=('Segoe UI', 15, "bold"), command=lambda title=text: show_scoring(title))
            show_scoring_btn.place(x=115, y=215)

            show_review_btn = Button(frame, width=11, text="Show Review", bg="pink", fg="black", font=('Segoe UI', 15, "bold"), command=lambda title=text: show_review(title))
            show_review_btn.place(x=115, y=270)

    
    #------------------------------------------  BUTTONS ---------------------------------------------#

    my_reviews_btn = Button(side_bar_options_frame, font=('Segoe UI', 18, "bold"), bg="#29BFCD", fg="black", text="My Reviews", width=17, border=3)
    my_reviews_btn.place(x=0, y=6)

    new_releases_btn = Button(side_bar_options_frame, font=('Segoe UI', 18, "bold"), bg="#236A43", fg="white", text="<Page in Progress>", width=17, border=3)
    new_releases_btn.place(x=0, y=66)

    calendar_btn = Button(side_bar_options_frame, font=('Segoe UI', 18, "bold"), bg="#236A43", fg="white", text="<Page in Progress>", width=17, border=3)
    calendar_btn.place(x=0, y=126)

    saved_btn = Button(side_bar_options_frame, font=('Segoe UI', 18, "bold"), bg="#236A43", fg="white", text="<Page in Progress>", width=17, border=3)
    saved_btn.place(x=0, y=186)

    separator_green = Button(sidebar_frame, font=('Segoe UI', 1, "bold"), bg="#1EFF83", fg="black", width=380)
    separator_green.place(x=0, y=555)

    separator_green2 = Button(separator, font=('Segoe UI', 1, "bold"), bg="#1EFF83", fg="black", width=1200)
    separator_green2.place(x=13, y=22)

    separator_lightgreen = Button(header_frame, font=('Segoe UI', 1, "bold"), bg="#1EFF83", width=1250)
    separator_lightgreen.place(x=0, y=0)

    separator_green3 = Button(header_frame, font=('Segoe UI', 1, "bold"), bg="#1EFF83", width=1250)
    separator_green3.place(x=0, y=15)

    separator_darkgreen = Button(sidebar_frame, font=('Segoe UI', 1, "bold"), bg="#236A43", fg="black", width=380)
    separator_darkgreen.place(x=0, y=575)

    logout_btn = Button(sidebar_frame, font=('Segoe UI', 18, "bold"), bg="#236A43", fg="white", text="Log Out", width=18, border=3, command=logout)
    logout_btn.place(x=0, y=630)

    settings_btn = Button(sidebar_frame, font=('Segoe UI', 18, "bold"), bg="#236A43", fg="white", text="Settings", width=18, border=3,command=settings)
    settings_btn.place(x=0, y=690)

    profile_info_btn = Button(sidebar_frame, font=('Segoe UI', 7, "bold"), bg="#FF9822", width=52, text=f"NO. RATINGS ---       USER ID: {username}")
    profile_info_btn.place(x=0, y=230)

    #------------------------------------------  LABELS ---------------------------------------------#

    task_bar_label = Label(taskbar_frame, font=('Segoe UI', 30, "bold"), fg="white", bg="#236A43", text="Game Critic Program")
    task_bar_label.place(x=680, y=5)

    header_label = Label(header_frame, font=('Segoe UI', 20, "bold"), fg="black", bg="#29BFCD", text="Saved Games / Game Library")

    header_label.place(x=430, y=21)

    #------------------------------------------  IMAGES ---------------------------------------------#

    profile = Label(sidebar_frame, image=root.photo1, border=0)
    profile.place(x=5, y=5)

    logo = Label(image_frame, image=root.photo2, border=0, background="#236A43")
    logo.place(x=70, y=10)

    root.mainloop()

main()
