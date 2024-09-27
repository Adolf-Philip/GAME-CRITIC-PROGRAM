

# IMPORTING LIBRARIES
from tkinter import *
from PIL import Image, ImageTk
from tkinter import ttk
from tkinter.ttk import Combobox
from tkinter import messagebox
from openpyxl import Workbook, load_workbook

# Setting String
review_content = ""

########################################################### CLASSES ########################################################################
#-------------------------------------------- GAME CLASS ---------------------------------------------------------#
class Game:

    #---------- Initialisation --------#
    def __init__(self, title, shrt_des, total_score, option1_score, option2_score, option3_score, option4_score, review):
        self.title = title
        self.short_description = shrt_des
        self.total_score = total_score
        self.option1_score = option1_score
        self.option2_score = option2_score
        self.option3_score = option3_score
        self.option4_score = option4_score
        self.review = review

    #---------- Method1: Convert Into Dictionary --------#
    def to_dict(self):
        return {
            "title": self.title,
            "short_description": self.short_description,
            "total_score": self.total_score,
            "option1_score": self.option1_score,
            "option2_score": self.option2_score,
            "option3_score": self.option3_score,
            "option4_score": self.option4_score,
            "review": self.review
        }

    #---------- Method2: Get Data From Dictionary --------#
    def from_dict(data):
        return Game(
            data["title"], data["short_description"], data["total_score"],
            data["option1_score"], data["option2_score"], data["option3_score"], data["option4_score"], data["review"]
        )

#-------------------------------------------- GAME MANAGER CLASS ---------------------------------------------------------#
class GameManager:

    #---------- Initialisation --------#
    def __init__(self, file_path):
        self.file_path = file_path
        self.games = self.load_games()

    #---------- Method1: Load Games --------#
    def load_games(self):
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            games = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) == 8:
                    game = Game(*row)
                    games.append(game)
                else:
                    print("Skipping Row")
            return games
        except FileNotFoundError:
            return []

    #---------- Method 2: Save Games in File --------#
    def save_games(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "Short Description", "Total Score", "Performance", "Visual Elements", "Audio Elements", "Gameplay", "Review"])
        for game in self.games:
            ws.append([
                game.title, game.short_description, game.total_score,
                game.option1_score, game.option2_score, game.option3_score, game.option4_score, game.review
            ])
        wb.save(self.file_path)

    #---------- Method3: Add Game to Excel File --------#
    def add_game(self, game):
        self.games.append(game)
        self.save_games()

    #---------- Method4: Edit Game --------#
    def edit_game(self, title, new_game):
        for idx, game in enumerate(self.games):
            if game.title == title:
                self.games[idx] = new_game
                self.save_games()
                return True
        return False

    #---------- Method5: Delete Game --------#
    def delete_game(self, title):
        self.games = [game for game in self.games if not (game.title == title)]
        self.save_games()

    #---------- Method6: Get Game from their title --------#
    def get_games_by_name(self):
        return [game.title for game in self.games]
    
############################################### FUNCTIONS ##########################################################################
#-------------------------------------------- CHECK FOR VALID REVIEW ---------------------------------------------------------#
def check_review(review_entry, min_length=72):

    #------------ Makes Sure review fits all requirements:  --------#
    review_text = review_entry.get("1.0", END).strip()  # Get the text from the Text widget

    keywords = ["performance", "visual", "audio", "gameplay", "overall"]

    if len(review_text) < min_length:
        messagebox.showerror(title="Error", message=(f"Error: Review must be at least {min_length} characters."))
        return False

    for keyword in keywords:
        if keyword not in review_text:
            messagebox.showerror(title="Error", message="Error: Review must include all keywords")
            return False

    messagebox.showinfo(title="Success", message="Review Saved")
    return True

#-------------------------------------------- ADD REVIEW ---------------------------------------------------------#
def add_review(add_review_btn):

    #------------ Check for Review and then add review if all requirements are filled  --------#
    def on_save():
        review_entry.lower()
        if check_review(review_entry):
            global review_content
            review_content = review_entry.get("1.0", END).strip()
            window.destroy()
            add_review_btn.configure(width=1, height=1,font=('Segoe UI', 1, "bold"))
            review_complete = Label(window_ref, text="Review Complete", font=('Segoe UI', 20, "bold"), fg="black")
            review_complete.place(x=80, y=645)

    #------------ Small Window for Adding a Review --------#
    window = Tk()
    window.title("Add Review Window")
    window.geometry("820x660")

    requirement_lbl = Label(window, text="Requirements for Review:\n- 72 characters minimum\n- All keywords must be present within the review\nKeywords: Performance, Visual Elements, Audio Elements, Gameplay and Overall Score", fg="black", justify="center", font=('Segoe UI', 15, "bold"))
    requirement_lbl.pack(side="top")
    review_entry = Text(window, font=('Segoe UI', 10, "bold"), width=480, height=25)
    review_entry.pack(side="top", pady=20, padx=10)

    complete_review_btn = Button(window, width=50, height=1, justify=CENTER, bg="light green", fg="black", text="Save Review", font=('Segoe UI', 20, "bold"), command=on_save)
    complete_review_btn.pack(side="top", padx=10, pady=0)

    window.mainloop()

#-------------------------------------------- ADD GAME ---------------------------------------------------------#
def add_games():


    manager = GameManager("game_review2.xlsx")


    #------------ For Completion of a game it adds to File --------#
    def complete_game():
        manager = GameManager("game_review2.xlsx")
        try:
            title = title_entry.get().strip()
            shrt_des = shrt_des_entry.get("1.0", END).strip()
            option1_score = int(score1_entry.get())
            option2_score = int(score2_entry.get())
            option3_score = int(score3_entry.get())
            option4_score = int(score4_entry.get())
            total_score = int(score5_entry.get())
            review = str(review_content)
            
        except ValueError:
            messagebox.showinfo("Error", "All fields are required")
            return

        if not title or not shrt_des or not review_content:
            messagebox.showerror("Error", "All fields are required.")
            return

        game = Game(title, shrt_des, total_score, option1_score, option2_score, option3_score, option4_score, review)
        manager.add_game(game)
        messagebox.showinfo("Success", f"Game '{title}' added successfully!")
        window.destroy()
    

    #------------ Window for Adding a Game --------#
    global window_ref
    window = Tk()
    window_ref = window
    window.title("Add Game Window")
    window.geometry("400x800")

    options = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
    total_score_options = [i for i in range(1, 101)]

    title_lbl = Label(window, text="Game Title:", font=('Segoe UI', 20, "bold"), fg="black")
    title_lbl.place(x=10, y=10)

    title_entry = Entry(window, width=13, justify=CENTER, font=('Segoe UI', 20, "bold"), fg="black")
    title_entry.place(x=170, y=10)

    shrt_des_lbl = Label(window, text="Description of game:", font=('Segoe UI', 20, "bold"), fg="black")
    shrt_des_lbl.place(x=60, y=75)

    shrt_des_entry = Text(window, width=53, height=10, font=('Segoe UI', 10, "bold"), fg="black")
    shrt_des_entry.place(x=10, y=125)

    scoring_frame = Frame(window, bg="light gray", height=300, width=375)
    scoring_frame.place(x=10, y=325)

    scoring_lbl = Label(scoring_frame, text="Scoring:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    scoring_lbl.place(x=130, y=10)

    performance_lbl = Label(scoring_frame, text="Performance:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    performance_lbl.place(x=10, y=80)

    score1_entry = Combobox(scoring_frame, values=options, font=('Segoe UI', 20, "bold"), width=5)
    score1_entry.place(x=250, y=80)

    visual_effects_lbl = Label(scoring_frame, text="Visual Effects:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    visual_effects_lbl.place(x=10, y=120)

    score2_entry = Combobox(scoring_frame, values=options, font=('Segoe UI', 20, "bold"), width=5)
    score2_entry.place(x=250, y=120)

    audio_effects_lbl = Label(scoring_frame, text="Audio Effects:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    audio_effects_lbl.place(x=10, y=160)

    score3_entry = Combobox(scoring_frame, values=options, font=('Segoe UI', 20, "bold"), width=5)
    score3_entry.place(x=250, y=160)

    gameplay_lbl = Label(scoring_frame, text="Gameplay:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    gameplay_lbl.place(x=10, y=200)

    score4_entry = Combobox(scoring_frame, values=options, font=('Segoe UI', 20, "bold"), width=5)
    score4_entry.place(x=250, y=200)

    total_score_lbl = Label(scoring_frame, text="Total Score:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    total_score_lbl.place(x=10, y=260)

    score5_entry = Combobox(scoring_frame, values=total_score_options, font=('Segoe UI', 20, "bold"), width=5)
    score5_entry.place(x=250, y=255)

    add_review_btn = Button(window, width=23, height=1, justify=CENTER, bg="light pink", fg="black", text="Add Review", font=('Segoe UI', 20, "bold"), command=lambda: add_review(add_review_btn))
    add_review_btn.place(x=10, y=635)

    #------------ Once a review is complete it will show that they have completed one--------#
    if review_content != "":
        add_review_btn.configure(width=1, height=1,font=('Segoe UI', 1, "bold"))
        review_complete = Label(window, text="Review Complete", font=('Segoe UI', 20, "bold"), fg="black")
        review_complete.place(x=80, y=645)

    complete_game_btn = Button(window, width=23, height=1, justify=CENTER, bg="light green", fg="black", text="Save Game", font=('Segoe UI', 20, "bold"), command=complete_game)
    complete_game_btn.place(x=10, y=710)

    window.mainloop()

#-------------------------------------------- EDIT GAME ---------------------------------------------------------#
def edit_games():

    manager = GameManager("game_review2.xlsx")

    #------------ Loading the game details of the one they have selected--------#
    def load_game_details(title):
        game = next((g for g in manager.games if g.title == title), None)

        if game:
            title_entry.delete(0, END)
            title_entry.insert(0, game.title)
            shrt_des_entry.delete("1.0", END)
            shrt_des_entry.insert("1.0", game.short_description)
            score1_entry.set(game.option1_score)
            score2_entry.set(game.option2_score)
            score3_entry.set(game.option3_score)
            score4_entry.set(game.option4_score)
            score5_entry.set(game.total_score)
            global review_content
            review_content = game.review

    #------------ Once they have completed editing, it will save their inputs --------#
    def complete_game():
        title = title_entry.get().strip()
        shrt_des = shrt_des_entry.get("1.0", END).strip()
        option1_score = int(score1_entry.get())
        option2_score = int(score2_entry.get())
        option3_score = int(score3_entry.get())
        option4_score = int(score4_entry.get())
        total_score = int(score5_entry.get())
        review = str(review_content)

        if not title or not shrt_des or not review_content:
            messagebox.showerror("Error", "All fields are required.")
            return

        new_game = Game(title, shrt_des, total_score, option1_score, option2_score, option3_score, option4_score, review)
        if manager.edit_game(selected_game.get(), new_game):
            messagebox.showinfo("Success", f"Game '{title}' edited successfully!")
            window.destroy()
        else:
            messagebox.showerror("Error", "Game not found.")

    #------------ Window for Editing a Game --------#
    global window_ref
    window = Tk()
    window_ref = window
    window.title("Edit Game Window")
    window.geometry("400x900")

    options = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
    total_score_options = [i for i in range(1, 101)]

    game_titles = manager.get_games_by_name()
    selected_game = StringVar(window)
    selected_game.set(game_titles[0])

    title_lbl = Label(window, text="Game Title:", font=('Segoe UI', 20, "bold"), fg="black")
    title_lbl.place(x=10, y=60)

    title_entry = Entry(window, width=13, justify=CENTER, font=('Segoe UI', 20, "bold"), fg="black")
    title_entry.place(x=170, y=60)

    shrt_des_lbl = Label(window, text="Description of game:", font=('Segoe UI', 20, "bold"), fg="black")
    shrt_des_lbl.place(x=60, y=125)

    shrt_des_entry = Text(window, width=53, height=10, font=('Segoe UI', 10, "bold"), fg="black")
    shrt_des_entry.place(x=10, y=175)

    scoring_frame = Frame(window, bg="light gray", height=300, width=375)
    scoring_frame.place(x=10, y=375)

    scoring_lbl = Label(scoring_frame, text="Scoring:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    scoring_lbl.place(x=130, y=10)

    performance_lbl = Label(scoring_frame, text="Performance:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    performance_lbl.place(x=10, y=80)

    score1_entry = Combobox(scoring_frame, values=options, font=('Segoe UI', 20, "bold"), width=5)
    score1_entry.place(x=250, y=80)

    visual_effects_lbl = Label(scoring_frame, text="Visual Effects:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    visual_effects_lbl.place(x=10, y=120)

    score2_entry = Combobox(scoring_frame, values=options, font=('Segoe UI', 20, "bold"), width=5)
    score2_entry.place(x=250, y=120)

    audio_effects_lbl = Label(scoring_frame, text="Audio Effects:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    audio_effects_lbl.place(x=10, y=160)

    score3_entry = Combobox(scoring_frame, values=options, font=('Segoe UI', 20, "bold"), width=5)
    score3_entry.place(x=250, y=160)

    gameplay_lbl = Label(scoring_frame, text="Gameplay:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    gameplay_lbl.place(x=10, y=200)

    score4_entry = Combobox(scoring_frame, values=options, font=('Segoe UI', 20, "bold"), width=5)
    score4_entry.place(x=250, y=200)

    total_score_lbl = Label(scoring_frame, text="Total Score:", font=('Segoe UI', 20, "bold"), fg="black", bg="light gray")
    total_score_lbl.place(x=10, y=260)

    score5_entry = Combobox(scoring_frame, values=total_score_options, font=('Segoe UI', 20, "bold"), width=5)
    score5_entry.place(x=250, y=255)

    add_review_btn = Button(window, width=23, height=1, justify=CENTER, bg="light pink", fg="black", text="Add Review", font=('Segoe UI', 20, "bold"), command=lambda: add_review(add_review_btn))
    add_review_btn.place(x=10, y=685)

    if review_content != "":
        add_review_btn.configure(width=1, height=1, font=('Segoe UI', 1, "bold"))
        review_complete = Label(window, text="Review Complete", font=('Segoe UI', 20, "bold"), fg="black")
        review_complete.place(x=80, y=685)

    complete_game_btn = Button(window, width=23, height=1, justify=CENTER, bg="light green", fg="black", text="Save Game", font=('Segoe UI', 20, "bold"), command=complete_game)
    complete_game_btn.place(x=10, y=760)

    game_select_lbl = Label(window, text="Select Game to Edit:", font=('Segoe UI', 20, "bold"), fg="black")
    game_select_lbl.place(x=10, y=10)

    game_select_menu = OptionMenu(window, selected_game, *game_titles, command=load_game_details)
    game_select_menu.place(x=280, y=20)

    load_game_details(selected_game.get())

    window.mainloop()

#-------------------------------------------- DELETE GAME ---------------------------------------------------------#
def delete_game():
    #------------ Window for deleting a game --------#
    window = Tk()
    window.title("Delete Game Window")
    window.geometry("400x400")
    

    #------------ Deleting the game they selected--------#
    def delete_selected_game():
        messagebox.showinfo(title="Message", message="Game Succefully Deleted")
        manager.delete_game(selected_game.get())
        window.destroy()


    manager = GameManager("game_review2.xlsx")

    game_titles = manager.get_games_by_name()
    selected_game = StringVar(window)
    selected_game.set(game_titles[0])


    game_select_lbl = Label(window, text="Select Game to Edit:", font=('Segoe UI', 20, "bold"), fg="black")
    game_select_lbl.place(x=10, y=10)

    game_select_menu = OptionMenu(window, selected_game, *game_titles)
    game_select_menu.place(x=280, y=20)

    delete_game_btn = Button(window,text="Delete Game", bg="red", font=('Segoe UI', 20, "bold"), fg="black", width=23, height=1, command=delete_selected_game)
    delete_game_btn.place(x=10, y= 60)

    window.mainloop()

#-------------------------------------------- DISPLAY GAMES ---------------------------------------------------------#
def display_games():

    manager = GameManager("game_review2.xlsx")

    # Get the list of game titles
    game_titles = manager.get_games_by_name()  

    window = Tk()
    window.title("All Games")
    window.geometry("800x600")  

    # Create a Canvas to allow for scrolling
    canvas = Canvas(window)
    canvas.pack(side=LEFT, fill=BOTH, expand=True)

    # Create a scrollbar
    scrollbar = Scrollbar(window, orient=VERTICAL, command=canvas.yview)
    scrollbar.pack(side=RIGHT, fill=Y)

    # Create a Frame to contain the game details
    frame = Frame(canvas)
    canvas.create_window((0, 0), window=frame, anchor="nw")

    # Update the scrollbar with the canvas
    frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    y = 10  # Initial Y position for the first game

    for title in game_titles:
        game = next((g for g in manager.games if g.title == title), None)  # Retrieve the Game object
        if game:
            details = (
                f"Title: {game.title}\n"
                f"Short Description: {game.short_description}\n"
                f"Performance Score: {game.option1_score}\n"
                f"Visual Score: {game.option2_score}\n"
                f"Audio Score: {game.option3_score}\n"
                f"Gameplay Score: {game.option4_score}\n"
                f"Total Score: {game.total_score}\n"
                f"Review: {game.review}\n"
                "\n\n"
            )
    
            game_label = Label(frame, text=details, font=('Segoe UI', 10, "bold"), fg="black", justify=LEFT)
            game_label.pack(anchor="w", padx=10, pady=10)

    window.mainloop()

############################################### MAIN PROGRAM: STARTING PAGE ##########################################################################
def main():
    root = Tk()
    root.title("Game Critic Program")
    root.geometry("400x300")

    header = Label(root, text="Game Critic Program", font=('Segoe UI', 20, "bold"), bg="white", fg="black")
    header.pack(side=TOP)

    # Create a 1x1 pixel image in memory
    pxl = Image.new('RGB', (1, 1), color='black')
    pxl = ImageTk.PhotoImage(pxl)

    add_btn = Button(root, text="Add Game", bg="light green", fg="black", justify=CENTER, image=pxl, compound='left', height=15, width=200, font=('Segoe UI', 20, "bold"), command=add_games)
    add_btn.pack(side=TOP, pady=5, ipady=15)

    edit_btn = Button(root, text="Edit Game", bg="orange", fg="black", image=pxl, justify=CENTER, compound='left', height=15, width=200, font=('Segoe UI', 20, "bold"), command=edit_games)
    edit_btn.pack(side=TOP, pady=5, ipady=15)

    delete_btn = Button(root, text="Delete Game", bg="red", fg="black", image=pxl, justify=CENTER, compound='left', height=15, width=200, font=('Segoe UI', 20, "bold"), command=delete_game)
    delete_btn.pack(side=TOP, pady=5, ipady=15)

    display_btn = Button(root, text="Display Game", bg="light blue", fg="black", image=pxl, justify=CENTER, compound='left', height=15, width=200, font=('Segoe UI', 20, "bold"), command=display_games)
    display_btn.pack(side=TOP, pady=5, ipady=15)

    root.mainloop()

main()
